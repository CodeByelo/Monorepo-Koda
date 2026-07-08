from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO
import unicodedata
import re
from pathlib import Path
from typing import Optional
import uuid

from database.async_db import get_db_connection
from auth.supabase_auth import get_current_user
from utils.idempotency import require_idempotency

router = APIRouter(prefix="/billing", tags=["billing"])


def validate_magic_bytes(file_bytes: bytes, filename: str) -> bool:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return file_bytes.startswith(b"%PDF")
    elif ext in (".xlsx", ".docx"):
        return file_bytes.startswith(b"PK\x03\x04")
    elif ext in (".xls", ".doc"):
        return file_bytes.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")
    elif ext == ".png":
        return file_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    elif ext in (".jpg", ".jpeg"):
        return file_bytes.startswith(b"\xff\xd8\xff")
    elif ext == ".webp":
        return len(file_bytes) >= 12 and file_bytes.startswith(b"RIFF") and file_bytes[8:12] == b"WEBP"
    return False


async def _log_security_event(
    conn,
    *,
    tenant_id: Optional[str],
    user_id: Optional[str],
    username: Optional[str],
    evento: str,
    detalles: Optional[str] = None,
    estado: str = "info",
    page: Optional[str] = None,
    ip_origen: Optional[str] = None,
    gerencia_id: Optional[int] = None,
) -> None:
    try:
        await conn.execute(
            """
            INSERT INTO security_events (tenant_id, user_id, username, evento, detalles, estado, page, ip_origen, gerencia_id, event_type)
            VALUES ($1::uuid, $2::uuid, $3, $4, $5, $6, $7, $8, $9, $10)
            """,
            uuid.UUID(str(tenant_id)) if tenant_id else None,
            uuid.UUID(str(user_id)) if user_id else None,
            username or "anon",
            evento,
            detalles,
            estado,
            page,
            ip_origen,
            gerencia_id,
            evento,
        )
    except Exception as e:
        import logging
        logging.getLogger("sistema_corporativo").warning(f"Error logging security event in billing: {e}")


def normalize_key(text: str) -> str:
    """Normaliza texto: minúsculas, sin acentos, sin caracteres especiales."""
    if not text:
        return ""
    text = str(text).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def find_col(headers_map: dict, *candidates) -> int | None:
    """Busca la columna probando múltiples nombres normalizados."""
    for candidate in candidates:
        key = normalize_key(candidate)
        if key in headers_map:
            return headers_map[key]
    return None


def safe_float(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_document_type(nota_credito: str, nota_debito: str, tipo_trans: str) -> str:
    if nota_credito:
        return "NC"
    if nota_debito:
        return "ND"
    if tipo_trans in ("03", "3"):
        return "NC"
    if tipo_trans in ("02", "2"):
        return "ND"
    return "NORMAL"


@router.post("/upload")
@require_idempotency
async def upload_billing_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    conn = Depends(get_db_connection)
):
    tenant_id = current_user.get("tenant_id")
    user_id = current_user.get("sub")
    username = current_user.get("username") or current_user.get("email")
    gerencia_id = current_user.get("gerencia_id")
    client_ip = request.client.host if request and request.client else None

    if tenant_id:
        await conn.execute("SELECT set_config('app.current_tenant_id', $1, true)", str(tenant_id))

    if not (file.filename or "").endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Formato de archivo no soportado")

    content = await file.read()
    if not validate_magic_bytes(content, file.filename):
        await _log_security_event(
            conn,
            tenant_id=tenant_id,
            user_id=user_id,
            username=username,
            evento="FILE_UPLOAD_SPOOFING",
            detalles=f"Intento de spoofing detectado al subir excel de facturación: {file.filename}. La firma binaria no coincide.",
            estado="critical",
            page="/billing/upload",
            ip_origen=client_ip,
            gerencia_id=gerencia_id
        )
        raise HTTPException(
            status_code=400,
            detail="El contenido del archivo no coincide con su extensión (se requiere un archivo Excel válido)."
        )
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active

    # Construir mapa de encabezados normalizados -> número de columna
    headers_map: dict[str, int] = {}
    raw_headers: list[str] = []
    for cell in ws[1]:
        raw = safe_str(cell.value)
        raw_headers.append(raw)
        key = normalize_key(raw)
        if key:
            headers_map[key] = cell.column

    # --- Mapeo flexible de columnas ---
    col_numero   = find_col(headers_map, "N° Op.", "No Op", "N Op", "Numero de Operacion", "num op", "nro op", "numero operacion", "N°")
    col_fecha    = find_col(headers_map, "Fecha del documento", "Fecha Documento", "Fecha", "fecha emision")
    col_rif      = find_col(headers_map, "N° R.I.F.", "Nro RIF", "RIF", "N° R.I.F", "NRO R.I.F", "rif contribuyente")
    col_nombre   = find_col(headers_map, "Nombre o Razon Social", "Razon Social", "Nombre", "nombre razon social", "razon social", "denominacion")
    col_factura  = find_col(headers_map, "N° de Factura", "Nro Factura", "Factura", "num factura", "numero factura")
    col_control  = find_col(headers_map, "N° Control del Documento", "Nro Control", "Control", "num control", "control documento", "numero control")
    col_nd       = find_col(headers_map, "N° de Nota de Debito", "Nota de Debito", "Nota Debito", "ND", "nro nota debito")
    col_nc       = find_col(headers_map, "N° de Nota de Credito", "Nota de Credito", "Nota Credito", "NC", "nro nota credito")
    col_tipo     = find_col(headers_map, "Tipo de Transaccion", "Tipo Transaccion", "Tipo Trans", "tipo", "tipo operacion")
    col_total    = find_col(headers_map, "Total Venta con IVA", "Total Venta", "Total", "total general", "monto total", "total bs")
    col_base16   = find_col(headers_map, "Base 16,00 %", "Base 16%", "Base Imponible 16", "base 16", "base imponible")
    col_iva16    = find_col(headers_map, "IVA 16,00 %", "IVA 16%", "IVA 16", "iva", "impuesto 16")

    def row_val(row, col):
        if col is None:
            return None
        return row[col - 1].value

    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Saltar filas completamente vacías
        if all(cell.value is None for cell in row):
            continue

        numero      = safe_str(row_val(row, col_numero))
        fecha       = safe_str(row_val(row, col_fecha))
        rif         = safe_str(row_val(row, col_rif))
        nombre      = safe_str(row_val(row, col_nombre))
        factura     = safe_str(row_val(row, col_factura))
        control     = safe_str(row_val(row, col_control))
        nota_debito = safe_str(row_val(row, col_nd))
        nota_credito= safe_str(row_val(row, col_nc))
        tipo_trans  = safe_str(row_val(row, col_tipo))
        total_venta = safe_float(row_val(row, col_total))
        base16      = safe_float(row_val(row, col_base16))
        iva16       = safe_float(row_val(row, col_iva16))

        tipo_doc = get_document_type(nota_credito, nota_debito, tipo_trans)

        total_calculado = base16 + iva16
        diff = abs(total_venta - total_calculado)
        errores = []
        cuadra = True
        if total_venta > 0 and diff > 0.5:
            cuadra = False
            errores.append(f"Diferencia de {diff:,.2f} entre Total Venta y Base+IVA calculados")

        data.append({
            "id": row_idx,
            "tipoDocumento": tipo_doc,
            "operacion": {
                "numero":              numero,
                "fechaDocumento":      fecha,
                "rif":                 rif,
                "nombreRazonSocial":   nombre,
                "factura":             factura,
                "controlDocumento":    control,
                "notaDebito":          nota_debito,
                "notaCredito":         nota_credito,
                "tipoTransaccion":     tipo_trans,
            },
            "contribuyente": [
                {"tasa": 16, "base": base16, "iva": iva16},
            ],
            "totales": {
                "baseTotal":       base16,
                "ivaTotal":        iva16,
                "totalCalculado":  total_calculado,
                "totalVenta":      total_venta,
            },
            "validacion": {
                "cuadra": cuadra,
                "errores": errores,
            },
        })

    # Devolver también las cabeceras detectadas para depuración
    return {
        "items": data,
        "_debug": {
            "headers_raw": raw_headers,
            "total_rows": len(data),
            "col_map": {
                "numero": col_numero, "rif": col_rif, "nombre": col_nombre,
                "total": col_total, "base16": col_base16, "iva16": col_iva16,
            }
        }
    }


@router.post("/export")
async def export_billing_excel(
    payload: dict,
    current_user: dict = Depends(get_current_user),
    conn = Depends(get_db_connection)
):
    tenant_id = current_user.get("tenant_id")
    if tenant_id:
        await conn.execute("SELECT set_config('app.current_tenant_id', $1, true)", str(tenant_id))
    items = payload.get("items", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturacion Final"

    headers = [
        "Nº Op.", "Fecha", "Nº R.I.F.", "Nombre / Razón Social",
        "Nº Factura", "Nº Control", "Tipo", "Total Venta", "Base 16%", "IVA 16%"
    ]
    ws.append(headers)

    for item in items:
        op = item.get("operacion", {})
        totales = item.get("totales", {})
        contrib = item.get("contribuyente", [{}])
        ws.append([
            op.get("numero"),
            op.get("fechaDocumento"),
            op.get("rif"),
            op.get("nombreRazonSocial"),
            op.get("factura"),
            op.get("controlDocumento"),
            item.get("tipoDocumento"),
            totales.get("totalVenta", 0),
            contrib[0].get("base", 0) if contrib else 0,
            contrib[0].get("iva", 0) if contrib else 0,
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Facturacion_Final.xlsx"},
    )
