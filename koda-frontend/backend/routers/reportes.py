from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List
from decimal import Decimal

from backend.core.database import get_db
from backend.services.reportes import ReporteService
from backend.services.auth import role_required
from backend.core.security import get_current_user

router = APIRouter(
    prefix="/reportes",
    tags=["Reportes Financieros"]
)

# --- Pydantic Schemas ---

class CuentaBalanceResponse(BaseModel):
    cuenta_codigo: str
    cuenta_nombre: str
    debe_usd: Decimal
    haber_usd: Decimal
    saldo_neto_usd: Decimal

    class Config:
        from_attributes = True

class BalanceComprobacionResponse(BaseModel):
    cuentas: List[CuentaBalanceResponse]
    total_debe_usd: Decimal
    total_haber_usd: Decimal
    cuadrado: bool

class EstadoResultadosResponse(BaseModel):
    ingresos_totales_usd: Decimal
    costos_totales_usd: Decimal
    gastos_totales_usd: Decimal
    utilidad_bruta_usd: Decimal
    utilidad_neta_usd: Decimal

class DashboardResumenResponse(BaseModel):
    saldo_bancos_usd: Decimal
    saldo_cxc_usd: Decimal

# --- Endpoints ---

@router.get("/balance-comprobacion", response_model=BalanceComprobacionResponse, dependencies=[Depends(role_required(['Admin', 'Contabilidad']))])
def get_balance_comprobacion(db: Session = Depends(get_db)):
    """
    Obtiene el balance de comprobación agrupado por cuenta contable,
    verificando la igualdad (Debe == Haber) general.
    """
    try:
        return ReporteService.obtener_balance_comprobacion(db)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar balance de comprobación: {str(e)}"
        )

@router.get("/estado-resultados", response_model=EstadoResultadosResponse, dependencies=[Depends(role_required(['Admin', 'Contabilidad']))])
def get_estado_resultados(db: Session = Depends(get_db)):
    """
    Obtiene el estado de resultados consolidado en USD (Ingresos, Costos, Gastos,
    Utilidad Bruta y Neta) a partir del libro contable.
    """
    try:
        return ReporteService.obtener_estado_resultados(db)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar estado de resultados: {str(e)}"
        )

@router.get("/dashboard-resumen", response_model=DashboardResumenResponse, dependencies=[Depends(role_required(['Admin']))])
def get_dashboard_resumen(db: Session = Depends(get_db)):
    """
    Obtiene métricas resumidas para el dashboard ejecutivo: saldo acumulado en Bancos
    y total pendiente por cobrar (CxC).
    """
    try:
        return ReporteService.dashboard_resumen(db)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar resumen del dashboard: {str(e)}"
        )

@router.get("/fiscal", dependencies=[Depends(role_required(['Admin', 'Contabilidad']))])
def get_reporte_fiscal(periodo: str, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """
    Obtiene el reporte de resumen del libro fiscal para un período dado.
    """
    from backend.models.operations import Venta
    from backend.models.erp_extended import Compra, RetencionIVA
    from backend.utils.helpers import to_float
    from sqlalchemy import extract
    
    try:
        y, m = map(int, periodo.split("-"))
    except:
        y, m = 2026, 7
        
    # Ventas del periodo
    ventas = db.query(Venta).filter(
        Venta.tenant_id == current_user.tenant_id,
        extract('year', Venta.fecha) == y,
        extract('month', Venta.fecha) == m,
        Venta.estado != "ANULADA"
    ).all()
    
    ventas_gravadas_base = sum(to_float(v.base_imponible) for v in ventas if v.tipo_factura != "EXENTA")
    ventas_gravadas_debito = sum(to_float(v.iva) for v in ventas if v.tipo_factura != "EXENTA")
    ventas_exoneradas = sum(to_float(v.base_imponible) for v in ventas if v.tipo_factura == "EXENTA")
    total_debitos = ventas_gravadas_debito
    
    # Compras del periodo
    compras = db.query(Compra).filter(
        Compra.tenant_id == current_user.tenant_id,
        extract('year', Compra.fecha) == y,
        extract('month', Compra.fecha) == m,
        Compra.estado != "ANULADA"
    ).all()
    
    compras_gravadas_base = sum(to_float(c.base_imponible) for c in compras if to_float(c.iva) > 0)
    compras_gravadas_credito = sum(to_float(c.iva) for c in compras if to_float(c.iva) > 0)
    compras_exentas = sum(to_float(c.base_imponible) for c in compras if to_float(c.iva) <= 0)
    total_creditos = compras_gravadas_credito
    
    # Retenciones
    retenciones = db.query(RetencionIVA).filter(
        RetencionIVA.tenant_id == current_user.tenant_id,
        RetencionIVA.periodo == periodo,
        RetencionIVA.tipo == "RECIBIDA"
    ).all()
    
    retenciones_soportadas = sum(to_float(r.monto_usd) for r in retenciones)
    
    cuota_tributaria = total_debitos - total_creditos - retenciones_soportadas
    
    return {
        "periodo": periodo,
        "ventas_gravadas_base": ventas_gravadas_base,
        "ventas_gravadas_debito": ventas_gravadas_debito,
        "ventas_exoneradas": ventas_exoneradas,
        "total_debitos": total_debitos,
        "compras_gravadas_base": compras_gravadas_base,
        "compras_gravadas_credito": compras_gravadas_credito,
        "compras_exentas": compras_exentas,
        "total_creditos": total_creditos,
        "retenciones_soportadas": retenciones_soportadas,
        "cuota_tributaria": cuota_tributaria
    }

@router.get("/fiscal/exportar", dependencies=[Depends(role_required(['Admin', 'Contabilidad']))])
def exportar_reporte_fiscal(periodo: str, formato: str, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """
    Exporta el reporte fiscal a PDF o Excel con datos reales del período.
    """
    from fastapi.responses import StreamingResponse
    import io

    data = get_reporte_fiscal(periodo, db, current_user)
    
    if formato == "pdf":
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0b5156'),
            alignment=1,
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#555555'),
            alignment=1,
            spaceAfter=25
        )
        
        story.append(Paragraph("KODA ERP - RESUMEN FISCAL (LIBRO DE IVA)", title_style))
        story.append(Paragraph(f"Periodo Fiscal: {periodo} | Documento Certificado DP-31", subtitle_style))
        
        table_data = [
            ['Concepto / Operación', 'Base Imponible (USD)', 'Débito / Crédito (USD)']
        ]
        
        # Sección Ventas (Débito)
        table_data.append(['VENTAS (DÉBITO FISCAL)', '', ''])
        table_data.append(['  Ventas Gravadas', f"{data['ventas_gravadas_base']:,.2f}", f"{data['ventas_gravadas_debito']:,.2f}"])
        table_data.append(['  Ventas Exoneradas / Exentas', f"{data['ventas_exoneradas']:,.2f}", "0.00"])
        table_data.append(['Total Débitos Fiscales (Ventas)', '', f"{data['total_debitos']:,.2f}"])
        
        # Sección Compras (Crédito)
        table_data.append(['COMPRAS (CRÉDITO FISCAL)', '', ''])
        table_data.append(['  Compras Gravadas', f"{data['compras_gravadas_base']:,.2f}", f"{data['compras_gravadas_credito']:,.2f}"])
        table_data.append(['  Compras Exentas', f"{data['compras_exentas']:,.2f}", "0.00"])
        table_data.append(['Total Créditos Fiscales (Compras)', '', f"{data['total_creditos']:,.2f}"])
        
        # Retenciones e Impuesto Neto
        table_data.append(['RESUMEN DE IMPUESTO', '', ''])
        table_data.append(['  (-) Retenciones de IVA Soportadas', '', f"{data['retenciones_soportadas']:,.2f}"])
        
        total_pago = data['cuota_tributaria']
        if total_pago < 0:
            table_data.append(['Excedente de Crédito Fiscal (S.F.)', '', f"{abs(total_pago):,.2f}"])
        else:
            table_data.append(['Total Cuota Tributaria a Pagar', '', f"{total_pago:,.2f}"])
            
        t = Table(table_data, colWidths=[280, 110, 110])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (2, 0), colors.HexColor('#0b5156')),
            ('TEXTCOLOR', (0, 0), (2, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (2, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (2, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            
            # Formatos de cabecera de secciones
            ('BACKGROUND', (0, 1), (2, 1), colors.HexColor('#e8f1f2')),
            ('FONTNAME', (0, 1), (2, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 5), (2, 5), colors.HexColor('#e8f1f2')),
            ('FONTNAME', (0, 5), (2, 5), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 9), (2, 9), colors.HexColor('#e8f1f2')),
            ('FONTNAME', (0, 9), (2, 9), 'Helvetica-Bold'),
            
            # Totales en negrita
            ('FONTNAME', (0, 4), (2, 4), 'Helvetica-Bold'),
            ('FONTNAME', (0, 8), (2, 8), 'Helvetica-Bold'),
            ('FONTNAME', (0, 11), (2, 11), 'Helvetica-Bold'),
        ]))
        
        story.append(t)
        
        # Certificaciones
        story.append(Spacer(1, 40))
        cert_data = [
            ['Preparado por: Contabilidad', 'Revisado por: Auditor Fiscal', 'Aprobado por: Representante Legal']
        ]
        ct = Table(cert_data, colWidths=[166, 166, 166])
        ct.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Oblique'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#555555')),
        ]))
        story.append(ct)
        
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=reporte_fiscal_{periodo}.pdf"}
        )
        
    elif formato == "excel":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen Fiscal"
        
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='0B5156', end_color='0B5156', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='116970', end_color='116970', fill_type='solid')
        section_font = Font(name='Arial', size=11, bold=True)
        total_font = Font(name='Arial', size=11, bold=True)
        
        ws.merge_cells('A1:C1')
        ws['A1'] = f"KODA ERP - RESUMEN FISCAL (LIBRO DE IVA) - PERIODO {periodo}"
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([])
        ws.append(['Concepto / Operación', 'Base Imponible (USD)', 'Débito / Crédito (USD)'])
        for col in ['A', 'B', 'C']:
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = header_fill
            
        ws.append(['VENTAS (DÉBITO FISCAL)'])
        ws.append(['  Ventas Gravadas', data['ventas_gravadas_base'], data['ventas_gravadas_debito']])
        ws.append(['  Ventas Exoneradas / Exentas', data['ventas_exoneradas'], 0.0])
        ws.append(['Total Débitos Fiscales (Ventas)', '', data['total_debitos']])
        
        ws.append(['COMPRAS (CRÉDITO FISCAL)'])
        ws.append(['  Compras Gravadas', data['compras_gravadas_base'], data['compras_gravadas_credito']])
        ws.append(['  Compras Exentas', data['compras_exentas'], 0.0])
        ws.append(['Total Créditos Fiscales (Compras)', '', data['total_creditos']])
        
        ws.append(['RESUMEN DE IMPUESTO'])
        ws.append(['  (-) Retenciones de IVA Soportadas', '', data['retenciones_soportadas']])
        
        total_pago = data['cuota_tributaria']
        if total_pago < 0:
            ws.append(['Excedente de Crédito Fiscal (S.F.)', '', abs(total_pago)])
        else:
            ws.append(['Total Cuota Tributaria a Pagar', '', total_pago])
            
        for row in [4, 8, 12]:
            ws[f'A{row}'].font = section_font
        for row in [7, 11, 14, 15]:
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = total_font
                
        for col in ['B', 'C']:
            for row in range(4, 16):
                cell = ws[f'{col}{row}']
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '$#,##0.00'
                    
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reporte_fiscal_{periodo}.xlsx"}
        )

    return {"ok": True}
