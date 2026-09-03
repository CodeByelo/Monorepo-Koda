from fastapi import APIRouter, Depends, Query, HTTPException, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date, timezone, timedelta
from decimal import Decimal
from typing import Optional, List

from backend.core.database import get_db
from backend.models.operations import (
    Venta, Cliente, Proveedor, Producto, VentaDetalle, KardexMovimiento, EvaluacionProveedor
)
from backend.models.erp_extended import (
    Compra, CuentaPorCobrar, CuentaPorPagar, CuentaBancaria, MovimientoBancario,
    Cotizacion, CotizacionItem, OrdenVenta, RequisicionCompra, TransferenciaInventario,
    RetencionIVA, RetencionISLR, Vendedor, Almacen, RecepcionStock, DevolucionProveedor, LoteProducto,
    NotaCredito, AnticipoCliente, Cheque, FondoCajaChica, GastoCajaChica, StockPorAlmacen,
    NotaEntrega, NotaEntregaItem, TransferenciaTesoreria, PrestamoUVC, PresupuestoPartida,
    ColocacionInversion, AuditoriaLog
)
from backend.schemas.operations import (
    CotizacionCreate, CotizacionStatusUpdate, CompraCreate, RecepcionStockCreate, RecepcionStockResponse,
    DevolucionProveedorCreate, NotaEntregaCreate, NotaEntregaEstadoUpdate
)
from backend.core.security import get_current_user, require_role
from backend.models.core import TasaCambio
from backend.utils.helpers import to_float, periodo_rango, ventas_periodo, tasa_actual, margen_bruto_pct, get_almacen_principal_id, verificar_periodo_abierto
from backend.services.contabilidad import ContabilidadService
from backend.routers.operaciones._shared import _as_aware, ISLR_WITHHOLDING_TABLE, _resolver_islr_automatico, calcular_reserva_fiscal

tesoreria_router = APIRouter(prefix="/tesoreria", tags=["Tesorería"], dependencies=[Depends(get_current_user)])


@tesoreria_router.get("/dashboard")
def tesoreria_dashboard(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    def format_currency(val):
        return f"${val:,.2f}"

    todas_cuentas = db.query(CuentaBancaria).filter(
        CuentaBancaria.activa == True,
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).all()
    
    bancos_bs = 0.0
    efectivo_zelle = 0.0
    custodia = 0.0
    bancos_lista = []
    
    for c in todas_cuentas:
        saldo = to_float(c.saldo_actual_usd)
        nombre = c.banco.lower()
        
        if "zelle" in nombre or "efectivo" in nombre:
            efectivo_zelle += saldo
        elif "custodia" in nombre:
            custodia += saldo * 0.99
        elif c.moneda == "VES":
            bancos_bs += saldo
        else:
            bancos_bs += saldo

        moneda = "VES" if c.moneda == "VED" else c.moneda
        bancos_lista.append({
            "nombre": c.banco,
            "saldo": format_currency(saldo),
            "neto": format_currency(saldo),
            "alerta": saldo < 100,
            "metadata": f"{moneda} - {c.numero_cuenta[-4:]}" if c.numero_cuenta else moneda,
            "icono": c.banco[0].upper() if c.banco else "B"
        })
    
    reserva_fiscal = calcular_reserva_fiscal(db, current_user.tenant_id)

    efectivo_transito = db.query(func.sum(MovimientoBancario.monto_usd)).filter(
        MovimientoBancario.estado == "PENDIENTE",
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).scalar()
    efectivo_transito = to_float(efectivo_transito)

    cheques_pd = db.query(func.sum(Cheque.monto_usd)).filter(
        Cheque.estado == "POST_DATADO",
        Cheque.tenant_id == current_user.tenant_id
    ).scalar()
    cheques_por_cobrar = to_float(cheques_pd)
    cheques_restar = 0.0 # Cheques emitidos (cxp) si hubiera una tabla, por ahora 0.
    
    # Cuarentena: sum of cuarentena_usd from all bank accounts
    cuarentena_restar = sum(to_float(c.cuarentena_usd) for c in todas_cuentas)

    disponibilidad_total = (bancos_bs + efectivo_zelle + custodia) - (cheques_restar + cuarentena_restar)
    liquidez_real = disponibilidad_total - reserva_fiscal

    ahora = datetime.now(timezone.utc)
    cxp_vencidas = db.query(CuentaPorPagar).filter(
        CuentaPorPagar.estado != "PAGADA",
        CuentaPorPagar.fecha_vencimiento < ahora,
        CuentaPorPagar.tenant_id == current_user.tenant_id
    ).all()
    
    # Proyeccion 7 dias
    limite_7d = ahora + timedelta(days=7)
    
    cxc_7d = db.query(CuentaPorCobrar).filter(
        CuentaPorCobrar.estado != "PAGADA",
        CuentaPorCobrar.fecha_vencimiento >= ahora,
        CuentaPorCobrar.fecha_vencimiento <= limite_7d,
        CuentaPorCobrar.tenant_id == current_user.tenant_id
    ).all()
    ingresos_7d = sum(to_float(c.monto_total_usd - c.monto_pagado_usd) for c in cxc_7d)

    cxp_7d = db.query(CuentaPorPagar).filter(
        CuentaPorPagar.estado != "PAGADA",
        CuentaPorPagar.fecha_vencimiento >= ahora,
        CuentaPorPagar.fecha_vencimiento <= limite_7d,
        CuentaPorPagar.tenant_id == current_user.tenant_id
    ).all()
    egresos_7d = sum(to_float(c.monto_total_usd - c.monto_pagado_usd) for c in cxp_7d)
    
    alertas_lista = []
    if cxp_vencidas:
        total_vencido = sum(to_float(c.monto_total_usd - c.monto_pagado_usd) for c in cxp_vencidas)
        alertas_lista.append({
            "tipo": "ALERTA",
            "titulo": f"{len(cxp_vencidas)} CXP VENCIDAS",
            "descripcion": f"Total: {format_currency(total_vencido)}",
            "bg": "bg-red-50",
            "color": "text-red-600"
        })
    
    if reserva_fiscal > 0:
        alertas_lista.append({
            "tipo": "FISCAL",
            "titulo": "DECLARACIÓN PENDIENTE",
            "descripcion": f"Retenciones retenidas: {format_currency(reserva_fiscal)}",
            "bg": "bg-amber-50",
            "color": "text-amber-600"
        })

    if not alertas_lista:
        alertas_lista.append({
            "tipo": "INFO",
            "titulo": "SIN COMPROMISOS URGENTES",
            "descripcion": "El flujo de caja está sano.",
            "bg": "bg-green-50",
            "color": "text-green-600"
        })

    return {
        "disponibilidad": {
            "total": format_currency(disponibilidad_total),
            "bancos_bs": format_currency(bancos_bs),
            "efectivo_zelle": format_currency(efectivo_zelle),
            "custodia": format_currency(custodia),
            "cheques_restar": format_currency(cheques_restar),
            "cuarentena_restar": format_currency(cuarentena_restar)
        },
        "cuarentena": format_currency(cuarentena_restar),
        "cheques_por_cobrar": format_currency(cheques_por_cobrar),
        "alertas": alertas_lista,
        "proyeccion_7d": {
            "ingresos_esperados": ingresos_7d,
            "egresos_esperados": egresos_7d
        },
        "metricas": [
            { "label": "Liquidez Real", "value": format_currency(liquidez_real), "desc": "Neto de compromisos", "color": "text-green-600", "border": "border-l-4 border-green-500" },
            { "label": "Cheques por Cobrar", "value": format_currency(cheques_por_cobrar), "desc": "Emitidos (Post-datados)", "color": "text-red-600", "border": "border-l-4 border-red-500" },
            { "label": "Efectivo en Tránsito", "value": format_currency(efectivo_transito), "desc": "Depósitos pendientes", "color": "text-amber-600", "border": "border-l-4 border-amber-500" },
            { "label": "Reserva Fiscal", "value": format_currency(reserva_fiscal), "desc": "IVA/ISLR Retenido", "color": "text-red-600", "border": "border-l-4 border-red-600" }
        ],
        "bancos": bancos_lista,
        "alertas": alertas_lista
    }


# ---- CUENTAS POR PAGAR (real, mirra la antigüedad de /cobranzas/* para CxC) ----

@tesoreria_router.get("/cuentas-por-pagar")
def cuentas_por_pagar_list(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Lista real de CuentaPorPagar del tenant, con antigüedad (días vencido)
    calculada sobre fecha_vencimiento vs. hoy, e IVA soportado tomado de la
    Compra vinculada (si existe). Mismo criterio de antigüedad que
    /cobranzas/cuentas usa para CuentaPorCobrar."""
    cxp = db.query(CuentaPorPagar).filter(
        CuentaPorPagar.tenant_id == current_user.tenant_id
    ).order_by(CuentaPorPagar.fecha_vencimiento.asc()).all()
    today = datetime.now(timezone.utc).date()

    compra_ids = [c.compra_id for c in cxp if c.compra_id]
    compras_map = {}
    if compra_ids:
        compras_map = {
            c.id: c for c in db.query(Compra).filter(Compra.id.in_(compra_ids)).all()
        }

    result = []
    for c in cxp:
        monto_total = to_float(c.monto_total_usd)
        monto_pagado = to_float(c.monto_pagado_usd)
        saldo = monto_total - monto_pagado
        vencimiento = c.fecha_vencimiento.date() if c.fecha_vencimiento else None
        dias_vencido = (today - vencimiento).days if vencimiento and today > vencimiento else 0
        estado_display = (
            "Vencida" if dias_vencido > 0
            else ("Por Vencer" if vencimiento and (vencimiento - today).days <= 7 else c.estado)
        )
        compra = compras_map.get(c.compra_id) if c.compra_id else None
        iva_soportado = to_float(compra.iva_usd) if compra else 0.0

        result.append({
            "id": c.id,
            "proveedor": c.proveedor.nombre if c.proveedor else "N/A",
            "rif": c.proveedor.rif if c.proveedor else "J-00000000-0",
            "numero_doc": c.numero_documento,
            "numero_control": compra.numero_control if compra and compra.numero_control else "N/A",
            "fecha_emision": c.fecha_emision.strftime("%d/%m/%Y") if c.fecha_emision else "",
            "fecha_vencimiento": vencimiento.strftime("%d/%m/%Y") if vencimiento else "",
            "monto_total": f"${monto_total:,.2f}",
            "monto_total_raw": monto_total,
            "monto_pagado": f"${monto_pagado:,.2f}",
            "saldo": f"${saldo:,.2f}",
            "saldo_raw": saldo,
            "iva_soportado": f"${iva_soportado:,.2f}",
            "iva_soportado_raw": iva_soportado,
            "dias_vencido": dias_vencido,
            "estado": estado_display,
            "estado_db": c.estado,
        })
    return result


@tesoreria_router.get("/cuentas-por-pagar/kpis")
def kpis_cuentas_por_pagar(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """KPIs reales para el header de Cuentas por Pagar (Tesorería)."""
    def format_currency(val):
        return f"${val:,.2f}"

    cxp = db.query(CuentaPorPagar).filter(
        CuentaPorPagar.tenant_id == current_user.tenant_id
    ).all()
    today = datetime.now(timezone.utc).date()
    abiertas = [c for c in cxp if c.estado != "PAGADA"]

    total_por_pagar = sum(to_float(c.monto_total_usd) - to_float(c.monto_pagado_usd) for c in abiertas)
    por_vencer_7d = sum(
        to_float(c.monto_total_usd) - to_float(c.monto_pagado_usd)
        for c in abiertas
        if c.fecha_vencimiento and 0 <= (c.fecha_vencimiento.date() - today).days <= 7
    )

    compra_ids = [c.compra_id for c in abiertas if c.compra_id]
    credito_fiscal_iva_usd = 0.0
    if compra_ids:
        compras = db.query(Compra).filter(Compra.id.in_(compra_ids)).all()
        credito_fiscal_iva_usd = sum(to_float(c.iva_usd) for c in compras)

    tasa = tasa_actual(db, current_user.tenant_id)
    credito_fiscal_iva_bs = credito_fiscal_iva_usd * tasa

    retenciones_pendientes = db.query(RetencionIVA).filter(
        RetencionIVA.tenant_id == current_user.tenant_id,
        RetencionIVA.tipo == "PRACTICADA",
        RetencionIVA.estado == "PENDIENTE",
    ).count()

    return {
        "metricas": [
            {"label": "Cuentas por Pagar", "value": format_currency(total_por_pagar), "desc": "Obligaciones totales", "color": "text-slate-800"},
            {"label": "A Pagar (7 días)", "value": format_currency(por_vencer_7d), "desc": "Requisición de flujo", "color": "text-amber-500"},
            {"label": "Crédito Fiscal IVA", "value": f"Bs. {credito_fiscal_iva_bs:,.2f}", "desc": "Acumulado a favor (CxP abiertas)", "color": "text-koda-main"},
            {"label": "Retenciones Pend.", "value": f"{retenciones_pendientes} Comp.", "desc": "Entregar a proveedores", "color": "text-blue-600"},
        ],
        "total_por_pagar_raw": total_por_pagar,
        "por_vencer_7d_raw": por_vencer_7d,
        "credito_fiscal_iva_raw": credito_fiscal_iva_bs,
        "retenciones_pendientes": retenciones_pendientes,
        "proveedores_vencidos": len({c.proveedor_id for c in abiertas if c.fecha_vencimiento and today > c.fecha_vencimiento.date()}),
    }


@tesoreria_router.get("/bancos")
def listar_bancos(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cuentas = db.query(CuentaBancaria).filter(CuentaBancaria.tenant_id == current_user.tenant_id).all()
    tasa = tasa_actual(db, current_user.tenant_id)
    res = []
    for c in cuentas:
        saldo_usd = to_float(c.saldo_actual_usd)
        saldo_local = saldo_usd * (tasa if c.moneda == "VES" else 1)
        movimientos_pendientes = db.query(MovimientoBancario).filter(
            MovimientoBancario.cuenta_id == c.id,
            MovimientoBancario.estado == "PENDIENTE"
        ).all()
        
        diferencia_pendiente = 0.00
        for m in movimientos_pendientes:
            monto_mov = to_float(m.monto_usd)
            if m.tipo == "EGRESO":
                monto_mov = -monto_mov
            diferencia_pendiente += monto_mov
        
        res.append({
            "id": c.id,
            "nombre": c.banco,
            "numero": c.numero_cuenta,
            "moneda": c.moneda,
            # Raw numbers for frontend calculation
            "saldo_contable_raw": saldo_local,
            "saldo_divisas_raw": saldo_usd,
            "diferencia_raw": diferencia_pendiente,
            # Formatted string fallbacks (optional, frontend should prefer formatting raw)
            "saldo_contable": f"{'Bs. ' if c.moneda == 'VES' else '$'}{saldo_local:,.2f}",
            "saldo_divisas": f"${saldo_usd:,.2f}",
            "diferencia": f"${diferencia_pendiente:,.2f}",
            "estado": "Activa" if c.activa else "Inactiva"
        })
    return res


@tesoreria_router.post("/bancos")
def crear_banco(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    c = CuentaBancaria(
        banco=body.get("nombre", ""),
        numero_cuenta=body.get("numero", ""),
        moneda=body.get("moneda", "VES"),
        saldo_actual_usd=body.get("saldo_actual", 0),
        activa=(body.get("estado") == "Activa"),
        tenant_id=current_user.tenant_id
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


@tesoreria_router.get("/movimientos")
def movimientos_banco(periodo: str, skip: int = 0, limit: int = 500, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    inicio, fin = periodo_rango(periodo)
    movs = db.query(MovimientoBancario).filter(
        MovimientoBancario.fecha >= inicio,
        MovimientoBancario.fecha < fin,
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).order_by(MovimientoBancario.fecha.desc()).offset(skip).limit(limit).all()
    return [{
        "id": f"MOV-{str(m.id).zfill(4)}", 
        "fecha": m.fecha.strftime("%d/%m/%Y"), 
        "concepto": m.concepto, 
        "monto": to_float(m.monto), 
        "tipo": m.tipo,
        "referencia": m.referencia,
        "estado": m.estado,
        "banco": m.cuenta.banco if m.cuenta else "Banco Asociado",
        "moneda": m.cuenta.moneda if m.cuenta else "USD"
    } for m in movs]


@tesoreria_router.post("/movimientos/importar")
def importar_movimientos():
    return {"ok": True}


@tesoreria_router.get("/conciliacion")
def conciliacion(periodo: str, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    inicio, fin = periodo_rango(periodo)
    movs = db.query(MovimientoBancario).filter(
        MovimientoBancario.fecha >= inicio,
        MovimientoBancario.fecha < fin,
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).all()
    
    movimientos = []
    cuentas_dict = {}
    for m in movs:
        banco_name = m.cuenta.banco if m.cuenta else "Banco Asociado"
        if m.cuenta_id not in cuentas_dict:
            cuentas_dict[m.cuenta_id] = {
                "cuenta": banco_name,
                "saldo_banco": to_float(m.cuenta.saldo_actual_usd) if m.cuenta else 0.0,
                "movs": []
            }
        
        cuentas_dict[m.cuenta_id]["movs"].append(m)
        
        movimientos.append({
            "id": f"MOV-{str(m.id).zfill(4)}",
            "fecha": m.fecha.strftime("%d/%m/%Y"),
            "banco": banco_name, 
            "referencia": m.referencia or "-",
            "monto": f"${to_float(m.monto_usd):,.2f}",
            "documento": f"KODA-{str(m.id).zfill(4)}",
            "estado": m.estado,
            "tipo": m.tipo.capitalize() if m.tipo else "Desconocido"
        })
        
    resumen_cuentas = []
    for c_id, data in cuentas_dict.items():
        saldo_banco = data["saldo_banco"]
        
        pendientes = sum([to_float(m.monto_usd) if m.tipo == "INGRESO" else -to_float(m.monto_usd) for m in data["movs"] if m.estado != "Conciliado" and m.estado != "CONCILIADO"])
        diferencia = pendientes
        saldo_koda = saldo_banco - diferencia
        
        estado = "Cuadra" if abs(diferencia) < 0.01 else "Diferencia"
        
        resumen_cuentas.append({
            "cuenta": data["cuenta"],
            "saldo_banco": f"${saldo_banco:,.2f}",
            "saldo_koda": f"${saldo_koda:,.2f}",
            "diferencia": f"${diferencia:,.2f}",
            "estado": estado
        })
    
    return {
        "metrics": {
            "movimientos_count": len(movs),
            "conciliados": len([m for m in movs if m.estado == 'CONCILIADO']),
            "pendientes": len([m for m in movs if m.estado != 'CONCILIADO']),
            "monto_x_conciliar": f"${sum([to_float(m.monto_usd) for m in movs if m.estado != 'CONCILIADO']):,.2f}"
        },
        "movimientos": movimientos,
        "resumen_cuentas": resumen_cuentas,
        "diferencias": []
    }


from pydantic import BaseModel
class RelacionarRequest(BaseModel):
    movimiento_id: str
    documento_id: str


@tesoreria_router.get("/conciliacion/pendientes")
def obtener_documentos_pendientes(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # Devuelve ventas pendientes a modo de ejemplo
    from backend.models.erp_extended import Venta
    ventas = db.query(Venta).filter(
        Venta.estado != 'ANULADA',
        Venta.tenant_id == current_user.tenant_id
    ).limit(10).all()
    docs = []
    for v in ventas:
        docs.append({
            "id": v.numero_factura,
            "label": f"{v.numero_factura} (Cliente)",
            "monto": f"${to_float(v.total_usd):,.2f}"
        })
    return docs

@tesoreria_router.post("/conciliacion/relacionar")
def relacionar_documento(payload: RelacionarRequest, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    # movimiento_id viene como MOV-123
    try:
        mov_id = int(payload.movimiento_id.replace("MOV-", ""))
        mov = db.query(MovimientoBancario).filter(
            MovimientoBancario.id == mov_id,
            MovimientoBancario.tenant_id == current_user.tenant_id
        ).first()
        if mov:
            mov.estado = "CONCILIADO"
            mov.documento_referencia = payload.documento_id
            db.commit()
    except:
        pass
    return {"ok": True}


@tesoreria_router.get("/flujo-caja")
def flujo_caja_tesoreria(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cxc_rows = db.query(CuentaPorCobrar).filter(
        CuentaPorCobrar.estado != "PAGADA",
        CuentaPorCobrar.tenant_id == current_user.tenant_id
    ).all()
    cxp_rows = db.query(CuentaPorPagar).filter(
        CuentaPorPagar.estado != "PAGADA",
        CuentaPorPagar.tenant_id == current_user.tenant_id
    ).all()
    proyecciones = []
    for r in cxc_rows:
        saldo = to_float(r.monto_total - r.monto_pagado)
        proyecciones.append({
            "date": r.fecha_vencimiento.strftime("%d/%m/%Y"),
            "fecha": r.fecha_vencimiento.strftime("%d/%m/%Y"),
            "concept": f"Cobranza {r.numero_documento}",
            "concepto": f"Cobranza {r.numero_documento}",
            "sub": r.cliente.nombre if r.cliente else "",
            "detalle": r.cliente.nombre if r.cliente else "",
            "area": "Cobranzas",
            "amount": saldo,
            "monto": saldo,
            "type": "Entrada",
            "tipo": "Entrada",
            "status": r.estado,
            "estado": r.estado,
            "statusColor": "bg-green-50 text-green-600",
        })
    for r in cxp_rows:
        saldo = to_float(r.monto_total - r.monto_pagado)
        vencida = _as_aware(r.fecha_vencimiento) < datetime.now(timezone.utc)
        proyecciones.append({
            "date": r.fecha_vencimiento.strftime("%d/%m/%Y"),
            "fecha": r.fecha_vencimiento.strftime("%d/%m/%Y"),
            "concept": f"Pago {r.numero_documento}",
            "concepto": f"Pago {r.numero_documento}",
            "sub": r.proveedor.nombre if r.proveedor else "",
            "detalle": r.proveedor.nombre if r.proveedor else "",
            "area": "Pagos",
            "amount": saldo,
            "monto": saldo,
            "type": "Salida",
            "tipo": "Salida",
            "status": r.estado,
            "estado": r.estado,
            "isCritical": vencida,
            "statusColor": "bg-red-50 text-red-600" if vencida else "bg-amber-50 text-amber-600",
        })
    return {"proyecciones": sorted(proyecciones, key=lambda x: x["date"])}


@tesoreria_router.post("/conciliacion/marcar")
def marcar_movimiento(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    mov_id = body.get("id") or body.get("movimiento_id")
    estado = body.get("estado", "CONCILIADO")
    
    mov = db.query(MovimientoBancario).filter(
        MovimientoBancario.id == mov_id,
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).first()
    if not mov:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    mov.estado = estado
    db.commit()
    return {"ok": True, "id": mov_id, "estado": estado}

@tesoreria_router.post("/conciliacion/cerrar")
def cerrar_conciliacion_periodo(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    periodo = body.get("periodo")
    if not periodo:
        raise HTTPException(status_code=400, detail="Periodo es requerido")
    
    inicio, fin = periodo_rango(periodo)
    movs = db.query(MovimientoBancario).filter(
        MovimientoBancario.fecha >= inicio, 
        MovimientoBancario.fecha < fin,
        MovimientoBancario.estado != "CERRADO",
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).all()
    
    count = 0
    for m in movs:
        m.estado = "CERRADO"
        count += 1
        
    db.commit()
    return {"ok": True, "cerrados": count, "mensaje": f"Periodo {periodo} cerrado exitosamente"}


@tesoreria_router.get("/caja-chica")
def caja_chica(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    fondos = db.query(FondoCajaChica).filter(FondoCajaChica.tenant_id == current_user.tenant_id).all()
    gastos = db.query(GastoCajaChica).join(FondoCajaChica).filter(
        FondoCajaChica.tenant_id == current_user.tenant_id
    ).order_by(GastoCajaChica.fecha.desc(), GastoCajaChica.id.desc()).limit(50).all()

    total_asignado = sum(to_float(f.asignado_usd) for f in fondos)
    total_disponible = sum(to_float(f.disponible_usd) for f in fondos)
    soportes_pendientes = db.query(GastoCajaChica).join(FondoCajaChica).filter(
        FondoCajaChica.tenant_id == current_user.tenant_id,
        GastoCajaChica.soporte == "Sin Soporte"
    ).count()
    reintegro_sugerido = total_asignado - total_disponible

    return {
        "metricas": {
            "fondo_asignado": f"${total_asignado:,.2f}",
            "saldo_disponible": f"${total_disponible:,.2f}",
            "soportes_pendientes": str(soportes_pendientes),
            "reintegro_sugerido": f"${max(0, reintegro_sugerido):,.2f}"
        },
        "fondos": [
            {
                "id": f"FD-{f.id:03d}",
                "nombre": f.nombre,
                "responsable": f.responsable,
                "asignado": f"${to_float(f.asignado_usd):,.2f}",
                "disponible": f"${to_float(f.disponible_usd):,.2f}",
                "estado": f.estado.capitalize(),
                "color": "bg-green-100 text-green-700" if f.estado == "ACTIVO" else "bg-amber-100 text-amber-700"
            } for f in fondos
        ],
        "gastos": [
            {
                "id": f"GC-{g.id:04d}",
                "referencia": f"GC-{g.id:04d}",
                "concepto": g.concepto,
                "fondo": g.fondo.nombre if g.fondo else "N/A",
                "monto": f"${to_float(g.monto_usd):,.2f}",
                "soporte": g.soporte,
                "fecha": g.fecha.strftime("%d/%m/%Y"),
                "estado": g.estado.capitalize(),
                "color": "bg-slate-100 text-slate-700" if g.estado == "PROCESADO" else "bg-amber-100 text-amber-700"
            } for g in gastos
        ]
    }


@tesoreria_router.post("/caja-chica/movimiento")
def movimiento_caja_chica(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    fondo_id_str = str(body.get("fondo_id", "")).strip()
    if not fondo_id_str:
        raise HTTPException(status_code=400, detail="fondo_id inválido o faltante")
    try:
        fid = int(fondo_id_str.replace("FD-", ""))
    except Exception:
        raise HTTPException(status_code=400, detail="fondo_id inválido o faltante")

    fondo = db.query(FondoCajaChica).filter(
        FondoCajaChica.id == fid,
        FondoCajaChica.tenant_id == current_user.tenant_id
    ).with_for_update().first()
    if not fondo:
        raise HTTPException(status_code=404, detail="Fondo de caja chica no encontrado")

    monto = float(body.get("monto", 0))
    fondo.disponible_usd = max(0, float(fondo.disponible_usd) - monto)

    gasto = GastoCajaChica(
        fondo_id=fid,
        concepto=body.get("concepto", "Gasto sin concepto"),
        monto_usd=monto,
        soporte=body.get("soporte", "Sin Soporte"),
        no_deducible=body.get("no_deducible", False),
        tenant_id=current_user.tenant_id
    )
    db.add(gasto)
    db.commit()
    return {"ok": True, "id": gasto.id}


@tesoreria_router.post("/caja-chica/reponer")
def reponer_caja_chica(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    fondos = db.query(FondoCajaChica).filter(FondoCajaChica.tenant_id == current_user.tenant_id).all()
    for f in fondos:
        f.disponible_usd = f.asignado_usd
    db.commit()
    return {"ok": True}


@tesoreria_router.post("/caja-chica/fondos")
def registrar_fondo_caja_chica(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    nombre = body.get("nombre", "")
    responsable = body.get("responsable", "")
    asignado_usd = float(body.get("asignado_usd", 0.0))
    
    if not nombre or not responsable:
        raise HTTPException(status_code=400, detail="Nombre y responsable requeridos")
        
    nuevo_fondo = FondoCajaChica(
        nombre=nombre,
        responsable=responsable,
        asignado_usd=asignado_usd,
        disponible_usd=asignado_usd,
        tenant_id=current_user.tenant_id
    )
    db.add(nuevo_fondo)
    db.commit()
    return {"ok": True, "id": nuevo_fondo.id}


@tesoreria_router.delete("/caja-chica/fondos/{fondo_id}")
def eliminar_fondo_caja_chica(fondo_id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    fondo = db.query(FondoCajaChica).filter(
        FondoCajaChica.id == fondo_id,
        FondoCajaChica.tenant_id == current_user.tenant_id
    ).first()
    if not fondo:
        raise HTTPException(status_code=404, detail="Fondo de caja chica no encontrado")

    tiene_gastos = db.query(GastoCajaChica).filter(GastoCajaChica.fondo_id == fondo.id).count() > 0
    nunca_usado = float(fondo.disponible_usd) == float(fondo.asignado_usd)

    if not tiene_gastos and nunca_usado:
        db.delete(fondo)
        db.commit()
        return {"ok": True, "accion": "eliminado"}
    else:
        # Tiene historial real (gastos registrados o saldo ya movido) — no se
        # borra para no perder trazabilidad financiera, se desactiva.
        fondo.estado = "INACTIVO"
        db.commit()
        return {"ok": True, "accion": "desactivado"}


@tesoreria_router.get("/arqueo")
def arqueo_caja(fecha: str, caja: str = "Caja Principal USD", db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """
    Calcula el saldo del sistema para el arqueo físico de caja.

    El saldo USD se calcula DINÁMICAMENTE sumando el total de todas las
    ventas ACTIVAS del día (fecha solicitada) del tenant, independientemente
    del método de pago. Esto refleja lo que realmente debería haber en caja
    según las operaciones del día.

    Anteriormente leía un saldo estático de CuentaBancaria que NO se
    actualizaba automáticamente con las ventas, causando saldos fantasma.
    """
    # --- Ventas reales del día por método de pago ---
    from sqlalchemy import cast, Date
    try:
        target_date = datetime.strptime(fecha, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        target_date = datetime.now(timezone.utc).date()

    ventas_del_dia = db.query(Venta).filter(
        Venta.estado == "ACTIVA",
        Venta.tenant_id == current_user.tenant_id,
        cast(Venta.fecha, Date) == target_date,
    ).all()

    # Desglose según método de pago:
    # Divisas / Efectivo USD vs Bolívares (Efectivo Bs, Pago Móvil, etc.)
    saldo_usd = 0.0
    saldo_ves = 0.0

    for v in ventas_del_dia:
        metodo = (v.metodo_pago or "").strip().upper()
        total_v_usd = to_float(v.total_usd or v.total or 0)
        tasa_v = to_float(v.tasa_cambio_bs) if to_float(v.tasa_cambio_bs) > 0 else 1.0

        if metodo in ("DIVISA", "DIVISAS", "USD", "DOLARES", "CASH_USD"):
            saldo_usd += total_v_usd
        elif metodo in ("EFECTIVO", "EFECTIVO_BS", "EFECTIVO BS", "VES", "BOLIVARES"):
            saldo_ves += total_v_usd * tasa_v
        else:
            # Otros métodos mixtos o generales: si la moneda de documento es SOLO_VES o el método es Bs
            if getattr(v, "moneda_documento", None) == "SOLO_VES":
                saldo_ves += total_v_usd * tasa_v
            else:
                saldo_usd += total_v_usd

    return {
        "fecha": fecha,
        "caja": caja,
        "saldo_sistema_usd": round(saldo_usd, 2),
        "saldo_sistema_ves": round(saldo_ves, 2),
        "ventas_count": len(ventas_del_dia),
    }


@tesoreria_router.post("/arqueo/cerrar")
def cerrar_arqueo(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    import json
    from backend.models.erp_extended import AuditoriaLog
    
    caja_name = body.get("caja", "Caja Principal USD")
    fisico_usd = float(body.get("fisico_usd", 0.0))
    fisico_ves = float(body.get("fisico_ves", 0.0))
    sistema_usd = float(body.get("sistema_usd", 0.0))
    cajero = body.get("cajero", "José Pérez")
    justificacion = body.get("justificacion", "")

    # Update balance of the selected cash account to reflect the physical count
    selected = db.query(CuentaBancaria).filter(
        CuentaBancaria.banco == caja_name,
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).first()
    if selected:
        selected.saldo_actual_usd = fisico_usd

    # Update balance of the VES cash account
    caja_ves = db.query(CuentaBancaria).filter(
        CuentaBancaria.banco == "Caja Principal VES",
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).first()
    if caja_ves:
        caja_ves.saldo_actual_usd = fisico_ves

    # Log the audit closure
    diferencia = fisico_usd - sistema_usd
    detalle_data = {
        "caja": caja_name,
        "sistema": sistema_usd,
        "fisico": fisico_usd,
        "diferencia": diferencia,
        "justificacion": justificacion,
        "resolucion": "Aceptable" if abs(diferencia) <= 50.0 else "Sujeto a Auditoría"
    }

    log_entry = AuditoriaLog(
        usuario=cajero,
        accion="CIERRE_ARQUEO",
        modulo="TESORERIA",
        detalle=json.dumps(detalle_data),
        fecha=datetime.now(timezone.utc),
        tenant_id=current_user.tenant_id
    )
    db.add(log_entry)
    db.commit()
    return {"ok": True}


@tesoreria_router.get("/arqueo/pdf")
def exportar_arqueo_pdf(
    fecha: str,
    caja: str = "Caja Principal USD",
    justificacion: str = "",
    fisico_ves: float = 0.0,
    denom_100: int = 0,
    denom_50: int = 0,
    denom_20: int = 0,
    denom_10: int = 0,
    denom_5: int = 0,
    denom_1: int = 0,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    from fastapi.responses import StreamingResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    # Retrieve system totals
    selected = db.query(CuentaBancaria).filter(
        CuentaBancaria.banco == caja,
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).first()
    system_usd = to_float(selected.saldo_actual_usd) if selected else 0.0

    caja_ves = db.query(CuentaBancaria).filter(
        CuentaBancaria.banco == "Caja Principal VES",
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).first()
    system_ves = to_float(caja_ves.saldo_actual_usd) if caja_ves else 0.0

    denoms = {
        "100": denom_100,
        "50": denom_50,
        "20": denom_20,
        "10": denom_10,
        "5": denom_5,
        "1": denom_1
    }
    fisico_usd = sum(int(k) * v for k, v in denoms.items())
    diff_usd = fisico_usd - system_usd
    diff_ves = fisico_ves - system_ves

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#0b5156'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#7f8c8d'),
        spaceAfter=25
    )
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0b5156'),
        spaceAfter=10
    )
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )

    story.append(Paragraph("ACTA DE ARQUEO FÍSICO DE CAJA", title_style))
    story.append(Paragraph(f"Fecha de Auditoría: {fecha} | Caja: {caja}", subtitle_style))
    story.append(Spacer(1, 10))

    # Summary Table
    data_summary = [
        [Paragraph("<b>CONCEPTO</b>", normal_style), Paragraph("<b>VALOR SISTEMA</b>", normal_style), Paragraph("<b>VALOR FÍSICO</b>", normal_style), Paragraph("<b>DIFERENCIA</b>", normal_style)],
        ["Efectivo USD", f"${system_usd:,.2f}", f"${fisico_usd:,.2f}", f"${diff_usd:,.2f}"],
        ["Efectivo VES", f"Bs. {system_ves:,.2f}", f"Bs. {fisico_ves:,.2f}", f"Bs. {diff_ves:,.2f}"]
    ]

    t_summary = Table(data_summary, colWidths=[150, 120, 120, 120])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f8f9fa')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#0b5156')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 20))

    # Denominations Table
    story.append(Paragraph("DESGLOSE DE EFECTIVO FÍSICO (USD)", section_style))
    data_denoms = [
        [Paragraph("<b>DENOMINACIÓN</b>", normal_style), Paragraph("<b>CANTIDAD</b>", normal_style), Paragraph("<b>TOTAL USD</b>", normal_style)],
    ]
    for denom, qty in sorted(denoms.items(), key=lambda x: int(x[0]), reverse=True):
        data_denoms.append([f"${denom}", str(qty), f"${int(denom) * qty:,.2f}"])
        
    t_denoms = Table(data_denoms, colWidths=[170, 170, 170])
    t_denoms.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f8f9fa')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(t_denoms)
    story.append(Spacer(1, 20))

    # Justification
    story.append(Paragraph("JUSTIFICACIÓN Y OBSERVACIONES", section_style))
    story.append(Paragraph(justificacion if justificacion else "SIN OBSERVACIONES DECLARADAS.", normal_style))
    story.append(Spacer(1, 40))

    # Signatures
    data_sigs = [
        ["_________________________", "_________________________"],
        ["Firma Responsable Caja", "Firma Auditor Autorizado"]
    ]
    t_sigs = Table(data_sigs, colWidths=[250, 250])
    t_sigs.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_sigs)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=acta_arqueo_{fecha}.pdf"}
    )


# ============================================================
# RUTAS MIGRADAS DESDE EXTRAS_EXT
# ============================================================

@tesoreria_router.get("/transferencias-internas")
def transferencias_internas(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    rows = db.query(TransferenciaTesoreria).filter(TransferenciaTesoreria.tenant_id == current_user.tenant_id).order_by(TransferenciaTesoreria.fecha.desc()).limit(30).all()
    return [
        {
            "id": f"TRF-{str(t.id).zfill(4)}", 
            "desc": "Transferencia" if t.tasa_cambio_bs == 1 else "Movimiento FX", 
            "from": t.origen.banco if t.origen else "Banco Origen",
            "to": t.destino.banco if t.destino else "Banco Destino",
            "amount": f"${to_float(t.monto_usd):,.2f}",
            "meta": f"Bs. {to_float(t.monto_usd * t.tasa_cambio_bs):,.2f} a Tasa {to_float(t.tasa_cambio_bs):,.4f}" if t.tasa_cambio_bs != 1 else "Misma moneda",
            "ref": t.concepto,
            "status": t.estado.capitalize(),
            "statusColor": "bg-green-100 text-green-700" if t.estado == "COMPLETADO" else ("bg-amber-100 text-amber-700" if t.estado == "PENDIENTE" else "bg-slate-100 text-slate-700"),
            "canConfirm": t.estado == "PENDIENTE",
            "fecha": t.fecha.strftime("%d/%m/%Y"),
            "db_id": t.id
        }
        for t in rows
    ]


@tesoreria_router.post("/transferencias-internas")
def registrar_transferencia(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    origen_id = body.get("origen_id")
    destino_id = body.get("destino_id")
    monto_usd = float(body.get("monto_usd", 0.0))
    tasa_cambio_bs = float(body.get("tasa_cambio_bs", 1.0))
    concepto = body.get("concepto", "Transferencia Interna")

    # Validar que ambas cuentas pertenezcan al tenant
    origen_ok = db.query(CuentaBancaria).filter(CuentaBancaria.id == origen_id, CuentaBancaria.tenant_id == current_user.tenant_id).first()
    destino_ok = db.query(CuentaBancaria).filter(CuentaBancaria.id == destino_id, CuentaBancaria.tenant_id == current_user.tenant_id).first()
    if not origen_ok or not destino_ok:
        raise HTTPException(status_code=400, detail="Cuentas bancarias no válidas o no pertenecen a su inquilino.")

    trf = TransferenciaTesoreria(
        cuenta_origen_id=origen_id,
        cuenta_destino_id=destino_id,
        monto_usd=monto_usd,
        tasa_cambio_bs=tasa_cambio_bs,
        concepto=concepto,
        estado="PENDIENTE",
        tenant_id=current_user.tenant_id
    )
    db.add(trf)
    db.commit()
    return {"ok": True, "id": trf.id}


@tesoreria_router.post("/transferencias-internas/{id}/confirmar")
def confirmar_transferencia(id: int, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    try:
        trf = db.query(TransferenciaTesoreria).filter(TransferenciaTesoreria.id == id, TransferenciaTesoreria.tenant_id == current_user.tenant_id).first()
        if not trf:
            raise HTTPException(status_code=404, detail="Transferencia no encontrada")
            
        if trf.estado == "COMPLETADO":
            return {"ok": True, "message": "Ya completada"}
            
        origen = db.query(CuentaBancaria).filter(CuentaBancaria.id == trf.cuenta_origen_id, CuentaBancaria.tenant_id == current_user.tenant_id).with_for_update().first()
        destino = db.query(CuentaBancaria).filter(CuentaBancaria.id == trf.cuenta_destino_id, CuentaBancaria.tenant_id == current_user.tenant_id).with_for_update().first()
        
        if not origen or not destino:
            raise HTTPException(status_code=400, detail="Cuentas no encontradas")

        monto_usd = to_float(trf.monto_usd)

        if origen:
            origen.saldo_actual_usd = to_float(origen.saldo_actual_usd) - monto_usd
        if destino:
            destino.saldo_actual_usd = to_float(destino.saldo_actual_usd) + monto_usd
            
        trf.estado = "COMPLETADO"

        # Generar Asiento Contable Automático de Transferencia Interna
        ContabilidadService.generar_asiento_transferencia_interna(
            monto=Decimal(str(trf.monto_usd)),
            tasa_cambio_bs=Decimal(str(trf.tasa_cambio_bs or 1.0)),
            concepto=f"Transferencia interna: {origen.banco} → {destino.banco}",
            referencia=f"TRF-{trf.id}",
            fecha=datetime.now(timezone.utc),
            db=db,
            tenant_id=current_user.tenant_id
        )

        db.commit()
        return {"ok": True}
    except HTTPException:
        db.rollback()
        raise


@tesoreria_router.get("/cuentas")
def obtener_cuentas(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cuentas = db.query(CuentaBancaria).filter(CuentaBancaria.activa == True, CuentaBancaria.tenant_id == current_user.tenant_id).all()
    return [
        {
            "id": c.id,
            "banco": c.banco,
            "numero_cuenta": c.numero_cuenta,
            "moneda": c.moneda,
            "saldo": to_float(c.saldo_actual_usd)
        }
        for c in cuentas
    ]


@tesoreria_router.get("/flujo")
def obtener_flujo_caja_alias(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cxc = db.query(CuentaPorCobrar).filter(CuentaPorCobrar.estado == "PENDIENTE", CuentaPorCobrar.tenant_id == current_user.tenant_id).all()
    cxp = db.query(CuentaPorPagar).filter(CuentaPorPagar.estado == "PENDIENTE", CuentaPorPagar.tenant_id == current_user.tenant_id).all()
    
    proyecciones = []
    
    for c in cxc:
        monto_pendiente = to_float(c.monto_total_usd) - to_float(c.monto_pagado_usd)
        if monto_pendiente > 0:
            proyecciones.append({
                "date": c.fecha_vencimiento.strftime("%d/%m/%Y") if c.fecha_vencimiento else "",
                "concept": f"Cobro Factura {c.numero_documento}",
                "sub": c.cliente.nombre if c.cliente else "Cliente General",
                "area": "Cobranzas",
                "amount": monto_pendiente,
                "type": "Entrada",
                "isCritical": False,
                "isBs": c.tasa_cambio_bs > 1.0,
                "status": "Pendiente",
                "statusColor": "bg-yellow-100 text-yellow-700"
            })
            
    for p in cxp:
        monto_pendiente = to_float(p.monto_total_usd) - to_float(p.monto_pagado_usd)
        if monto_pendiente > 0:
            is_critical = monto_pendiente > 500.0
            proyecciones.append({
                "date": p.fecha_vencimiento.strftime("%d/%m/%Y") if p.fecha_vencimiento else "",
                "concept": f"Pago Factura {p.numero_documento}",
                "sub": p.proveedor.nombre if p.proveedor else "Proveedor General",
                "area": "Compras",
                "amount": monto_pendiente,
                "type": "Salida",
                "isCritical": is_critical,
                "isBs": p.tasa_cambio_bs > 1.0,
                "status": "Pendiente",
                "statusColor": "bg-red-100 text-red-700" if is_critical else "bg-slate-100 text-slate-700"
            })
            
    try:
        proyecciones.sort(key=lambda x: datetime.strptime(x["date"], "%d/%m/%Y"))
    except Exception:
        pass
        
    return {
        "proyecciones": proyecciones
    }


@tesoreria_router.get("/turnos")
def obtener_auditoria_turnos(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    import json
    
    logs = db.query(AuditoriaLog).filter(AuditoriaLog.accion == "CIERRE_ARQUEO", AuditoriaLog.tenant_id == current_user.tenant_id).all()
    
    cajeros_monitoreados = set()
    desviacion_total = 0.0
    cajeros_con_alertas = set()
    
    historial_por_cajero = {}
    ranking_data = {}
    
    for log in logs:
        cajero = log.usuario
        cajeros_monitoreados.add(cajero)
        
        try:
            detalle = json.loads(log.detalle)
        except Exception:
            continue
            
        diff = to_float(detalle.get("diferencia", 0.0))
        caja = detalle.get("caja", "Caja General")
        fisico = to_float(detalle.get("fisico", 0.0))
        resolucion = detalle.get("resolucion", "Aceptable")
        
        desviacion_total += abs(diff)
        
        if cajero not in ranking_data:
            ranking_data[cajero] = {
                "name": cajero,
                "loss": 0.0,
                "role": "Cajero",
                "initials": "".join([part[0] for part in cajero.split() if part][:2]).upper()
            }
        ranking_data[cajero]["loss"] += diff
        
        if cajero not in historial_por_cajero:
            historial_por_cajero[cajero] = []
        historial_por_cajero[cajero].append({
            "date": log.fecha.strftime("%d/%m/%Y %H:%M"),
            "box": caja,
            "physical": f"${fisico:,.2f}",
            "diff": f"{'+' if diff >= 0 else ''}${diff:,.2f}",
            "resolution": resolucion
        })
        
    for cajero, r in ranking_data.items():
        if abs(r["loss"]) > 20.0:
            cajeros_con_alertas.add(cajero)
            
    ranking_list = []
    for c, r in ranking_data.items():
        loss_val = r["loss"]
        ranking_list.append({
            "name": r["name"],
            "role": r["role"],
            "initials": r["initials"],
            "loss": f"{'-' if loss_val < 0 else ''}${abs(loss_val):,.2f}",
            "desc": "Pérdida acumulada" if loss_val < 0 else "Sobrante acumulado",
            "isCritical": abs(loss_val) > 20.0
        })
        
    return {
        "metricas": {
            "cajeros_monitoreados": f"{len(cajeros_monitoreados)} Usuarios",
            "desviacion_total": f"${desviacion_total:,.2f}",
            "alertas_criticas": f"{len(cajeros_con_alertas)} Usuarios"
        },
        "ranking": ranking_list,
        "historiales": historial_por_cajero
    }


@tesoreria_router.get("/prestamos/resumen")
def resumen_prestamos_uvc(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    loans = db.query(PrestamoUVC).filter(PrestamoUVC.tenant_id == current_user.tenant_id).all()

    tasa_uvc_hoy = to_float(tasa_actual(db, current_user.tenant_id))
    tasas_recientes = (
        db.query(TasaCambio)
        .filter(
            (TasaCambio.tenant_id == current_user.tenant_id) | (TasaCambio.tenant_id.is_(None))
        )
        .order_by(TasaCambio.fecha.desc())
        .limit(2)
        .all()
    )
    tasa_uvc_ayer = to_float(tasas_recientes[1].valor_ves) if len(tasas_recientes) > 1 else tasa_uvc_hoy
    var_24h = ((tasa_uvc_hoy - tasa_uvc_ayer) / tasa_uvc_ayer) * 100 if tasa_uvc_ayer else 0.0
    
    total_uvc = 0.0
    total_reval_bs = 0.0
    
    loans_list = []
    for l in loans:
        monto_uvc_val = to_float(l.monto_uvc)
        tasa_inicial = to_float(l.tasa_cambio_bs)
        
        total_uvc += monto_uvc_val
        saldo_bs = monto_uvc_val * tasa_uvc_hoy
        reval_diff = (tasa_uvc_hoy - tasa_inicial) * monto_uvc_val
        total_reval_bs += reval_diff
        
        banks = ["Banesco", "Banco Provincial", "Banco de Venezuela", "Banco Mercantil"]
        bank_name = banks[l.id % len(banks)]
        
        loans_list.append({
            "id": l.id,
            "ref": f"CRE-UVC-{l.id:04d}",
            "descripcion": l.descripcion,
            "bank": bank_name,
            "capital": f"{monto_uvc_val:,.2f} UVC",
            "initRate": f"Bs. {tasa_inicial:,.2f}",
            "currentRate": f"Bs. {tasa_uvc_hoy:,.2f}",
            "balance": f"Bs. {saldo_bs:,.2f}",
            "status": l.estado,
            "color": "bg-green-100 text-green-700" if l.estado == "ACTIVO" else "bg-slate-100 text-slate-700"
        })
        
    return {
        "metricas": {
            "tasa_uvc_hoy": f"Bs. {tasa_uvc_hoy:,.2f}",
            "var_24h": f"{var_24h:+.2f}%",
            "capital_pendiente_uvc": f"{total_uvc:,.2f} UVC",
            "eqv_bs": f"Bs. {(total_uvc * tasa_uvc_hoy):,.2f}",
            "diff_indexacion": f"Bs. {total_reval_bs:,.2f}",
            "reval_desc": "Pérdida por revalorización mes" if total_reval_bs >= 0 else "Ganancia cambiaria acumulada"
        },
        "creditos": loans_list
    }


@tesoreria_router.post("/prestamos-uvc")
def registrar_prestamo_uvc(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    desc = body.get("descripcion", "Préstamo Comercial UVC")
    monto_uvc = float(body.get("monto_uvc", 0.0))
    tasa = float(body.get("tasa", 12.0))
    tasa_ref = to_float(tasa_actual(db, current_user.tenant_id))
    tasa_cambio_bs = float(body.get("tasa_cambio_bs", tasa_ref))

    saldo_usd = (monto_uvc * tasa_cambio_bs) / tasa_ref if tasa_ref else 0.0
    
    nuevo_prestamo = PrestamoUVC(
        descripcion=desc,
        monto_uvc=monto_uvc,
        tasa=tasa,
        saldo_usd=saldo_usd,
        tasa_cambio_bs=tasa_cambio_bs,
        estado="ACTIVO",
        fecha_inicio=datetime.now(timezone.utc),
        tenant_id=current_user.tenant_id
    )
    db.add(nuevo_prestamo)
    db.commit()
    return {"ok": True}


@tesoreria_router.get("/presupuesto")
def presupuesto_tesoreria(periodo: str = None, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    periodo = periodo or datetime.now(timezone.utc).strftime("%Y-%m")
    partidas = db.query(PresupuestoPartida).filter(PresupuestoPartida.periodo == periodo, PresupuestoPartida.tenant_id == current_user.tenant_id).all()
    return {"periodo": periodo, "partidas": [
        {"centro": p.centro_costo, "concepto": p.concepto, "presupuestado": to_float(p.presupuestado_usd), "ejecutado": to_float(p.ejecutado_usd)}
        for p in partidas
    ]}


@tesoreria_router.get("/presupuesto/desviacion")
def desviacion_presupuestaria(periodo: str = None, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    periodo = periodo or datetime.now(timezone.utc).strftime("%Y-%m")
    partidas = db.query(PresupuestoPartida).filter(PresupuestoPartida.periodo == periodo, PresupuestoPartida.tenant_id == current_user.tenant_id).all()
    
    oldest_mov = db.query(MovimientoBancario).filter(MovimientoBancario.tenant_id == current_user.tenant_id).order_by(MovimientoBancario.fecha.asc()).first()
    tasa_plan = to_float(oldest_mov.tasa_cambio_bs) if oldest_mov else to_float(tasa_actual(db, current_user.tenant_id))

    newest_mov = db.query(MovimientoBancario).filter(MovimientoBancario.tenant_id == current_user.tenant_id).order_by(MovimientoBancario.fecha.desc()).first()
    tasa_real_raw = to_float(newest_mov.tasa_cambio_bs) if newest_mov else to_float(tasa_actual(db, current_user.tenant_id))
    tasa_real = tasa_real_raw if tasa_real_raw > 5.0 else tasa_plan * 1.16
    
    total_plan_usd = 0.0
    total_real_usd = 0.0
    total_fx_impact_usd = 0.0
    total_inefficiency_usd = 0.0
    
    breakdown_list = []
    for p in partidas:
        plan_usd = to_float(p.presupuestado_usd)
        real_usd = to_float(p.ejecutado_usd)
        
        plan_bs = plan_usd * tasa_plan
        real_bs = real_usd * tasa_real
        
        total_plan_usd += plan_usd
        total_real_usd += real_usd
        
        deviation_usd = real_usd - plan_usd
        
        fx_impact_usd = 0.0
        inefficiency_usd = 0.0
        
        if deviation_usd > 0:
            fx_impact_usd = real_usd * (1 - (tasa_plan / tasa_real))
            inefficiency_usd = max(0.0, deviation_usd - fx_impact_usd)
            
            total_fx_impact_usd += fx_impact_usd
            total_inefficiency_usd += inefficiency_usd
            
        impact_str = f"-${deviation_usd:,.2f}" if deviation_usd > 0 else f"+${abs(deviation_usd):,.2f}"
        
        is_over = real_usd > plan_usd
        
        breakdown_list.append({
            "item": p.concepto,
            "centro": p.centro_costo,
            "planBs": f"{plan_bs:,.2f}",
            "realBs": f"{real_bs:,.2f}",
            "planUsd": f"{plan_usd:,.2f}",
            "realUsd": f"{real_usd:,.2f}",
            "impact": impact_str,
            "isOver": is_over,
            "status": "Sobregiro" if is_over else "Conforme",
            "cause": "Devaluación" if fx_impact_usd > inefficiency_usd else "Precios/Gestión",
            "statusColor": "bg-red-100 text-red-700" if is_over else "bg-green-100 text-green-700",
            "causeColor": "bg-red-50 text-red-600 border-red-100" if is_over else "bg-green-50 text-green-600 border-green-100"
        })
        
    total_deviation = total_real_usd - total_plan_usd
    fx_percent = (total_fx_impact_usd / total_real_usd * 100) if total_real_usd > 0 else 0.0
    ineff_percent = (total_inefficiency_usd / total_real_usd * 100) if total_real_usd > 0 else 0.0
    
    return {
        "periodo": periodo,
        "tasa_plan": f"Bs. {tasa_plan:,.2f}",
        "tasa_real": f"Bs. {tasa_real:,.2f}",
        "metricas": {
            "desviacion_total": f"${total_deviation:,.2f}" if total_deviation >= 0 else f"+${abs(total_deviation):,.2f}",
            "impacto_cambiario": f"{fx_percent:.1f}%",
            "ineficiencia_operativa": f"{ineff_percent:.1f}%"
        },
        "breakdown": breakdown_list,
        "distribucion": {
            "fx_pct": fx_percent,
            "ineff_pct": ineff_percent
        }
    }


@tesoreria_router.get("/inversiones/resumen")
def resumen_inversiones(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    placements = db.query(ColocacionInversion).filter(ColocacionInversion.estado == "ACTIVO", ColocacionInversion.tenant_id == current_user.tenant_id).all()

    tasa_real = to_float(tasa_actual(db, current_user.tenant_id))
    
    total_gain_bs = 0.0
    total_capital_bs = 0.0
    total_net_real_usd = 0.0
    
    placements_list = []
    for p in placements:
        cap_bs = to_float(p.capital_bs)
        rate_anual = to_float(p.tasa_interes_anual)
        plazo = p.plazo_dias
        init_rate = to_float(p.tasa_cambio_inicial)
        
        interest_bs = cap_bs * (rate_anual / 100.0) * (plazo / 360.0)
        total_gain_bs += interest_bs
        total_capital_bs += cap_bs
        
        cap_usd = cap_bs / init_rate if init_rate > 0 else 0.0
        final_usd = (cap_bs + interest_bs) / tasa_real
        real_result_usd = final_usd - cap_usd
        total_net_real_usd += real_result_usd
        
        fx_effect_bs = cap_bs * (1 - (init_rate / tasa_real)) if tasa_real else 0.0
        
        placements_list.append({
            "id": p.id,
            "name": p.nombre,
            "term": f"{plazo} Días",
            "capital": f"Bs. {cap_bs:,.2f}",
            "rates": f"{rate_anual}% Anual",
            "gain": f"Bs. {interest_bs:,.2f}",
            "fxEffect": f"-Bs. {fx_effect_bs:,.2f}",
            "realRes": f"${real_result_usd:,.2f}" if real_result_usd >= 0 else f"-${abs(real_result_usd):,.2f}",
            "isNegative": real_result_usd < 0
        })
        
    avg_interest = sum(to_float(p.tasa_interes_anual) for p in placements) / len(placements) if placements else 0.0
    bcv_dev_pct = 15.7
    
    eff_real_pct = (total_net_real_usd / (total_capital_bs / tasa_real) * 100) if total_capital_bs > 0 and tasa_real else 0.0
    
    return {
        "metricas": {
            "eficiencia_real": f"{eff_real_pct:.1f}%",
            "interes_acumulado": f"Bs. {total_gain_bs:,.2f}",
            "devaluacion_periodo": f"{bcv_dev_pct:.1f}%"
        },
        "colocaciones": placements_list,
        "interes_promedio": avg_interest,
        "devaluacion_bcv": bcv_dev_pct
    }


@tesoreria_router.post("/inversiones")
def registrar_inversion(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    nombre = body.get("nombre", "Colocación Plazo Fijo")
    plazo_dias = int(body.get("plazo_dias", 30))
    capital_bs = float(body.get("capital_bs", 0.0))
    tasa_interes = float(body.get("tasa_interes_anual", 48.0))
    tasa_cambio = float(body.get("tasa_cambio_inicial", 42.15))
    if tasa_cambio <= 0:
        raise HTTPException(status_code=400, detail="La tasa de cambio inicial debe ser mayor a 0")
    
    nueva_inv = ColocacionInversion(
        nombre=nombre,
        plazo_dias=plazo_dias,
        capital_bs=capital_bs,
        tasa_interes_anual=tasa_interes,
        tasa_cambio_inicial=tasa_cambio,
        fecha_inicio=datetime.now(timezone.utc),
        estado="ACTIVO",
        tenant_id=current_user.tenant_id
    )
    db.add(nueva_inv)
    db.commit()
    return {"ok": True}


@tesoreria_router.post("/importar")
def importar_extracto_bancario(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    """Importa un extracto bancario (filas ya parseadas en el cliente, ver
    ImportStatement.tsx) y es el mecanismo REAL de conciliación: cada fila del
    extracto se intenta cruzar contra un movimiento interno pendiente
    (creado por pagos, compras, etc.) por monto + fecha + tipo (reforzado por
    referencia cuando ambas la traen). Sólo si hay un match real se marca ese
    movimiento como CONCILIADO; si ninguna fila coincide no se marca nada en
    masa. Las filas del extracto sin contraparte interna se insertan como
    movimientos nuevos en estado ACTIVO, quedando disponibles para conciliar
    manualmente (endpoints /conciliacion/relacionar y /conciliacion/marcar)."""
    cuenta_id = body.get("cuenta_id")
    movs = body.get("movimientos", [])

    cuenta = db.query(CuentaBancaria).filter(CuentaBancaria.id == cuenta_id, CuentaBancaria.tenant_id == current_user.tenant_id).with_for_update().first()
    if not cuenta:
        return {"ok": False, "message": "Cuenta bancaria no encontrada"}

    tasa_cambio = to_float(tasa_actual(db, current_user.tenant_id)) or 36.42

    candidatos_pendientes = db.query(MovimientoBancario).filter(
        MovimientoBancario.cuenta_id == cuenta_id,
        MovimientoBancario.tenant_id == current_user.tenant_id,
        MovimientoBancario.estado != "CONCILIADO"
    ).all()

    TOLERANCIA_USD = 0.02
    TOLERANCIA_DIAS = 3

    total_monto_usd = 0.0
    conciliados_count = 0
    nuevos_count = 0

    for m in movs:
        fecha_str = m.get("fecha", datetime.now(timezone.utc).strftime("%Y-%m-%d"))
        ref = str(m.get("referencia") or "").strip()
        concepto = m.get("concepto", "Movimiento de extracto importado")
        monto_val = float(m.get("monto", 0.0))

        try:
            fecha_mov = datetime.strptime(fecha_str, "%Y-%m-%d") if "-" in fecha_str else datetime.now()
        except Exception:
            fecha_mov = datetime.now()

        monto_usd = monto_val / tasa_cambio if tasa_cambio else 0.0
        tipo = "INGRESO" if monto_usd >= 0 else "EGRESO"

        match = None
        for cand in candidatos_pendientes:
            if cand.tipo != tipo:
                continue
            if abs(to_float(cand.monto_usd) - abs(monto_usd)) > TOLERANCIA_USD:
                continue
            if cand.fecha and abs((cand.fecha - fecha_mov).days) > TOLERANCIA_DIAS:
                continue
            if ref and cand.referencia and ref.lower() != cand.referencia.strip().lower():
                continue
            match = cand
            break

        if match:
            match.estado = "CONCILIADO"
            if ref and not match.referencia:
                match.referencia = ref[:100]
            candidatos_pendientes.remove(match)
            conciliados_count += 1
        else:
            nuevo_mov = MovimientoBancario(
                cuenta_id=cuenta_id,
                fecha=fecha_mov,
                concepto=concepto,
                monto_usd=abs(monto_usd),
                tasa_cambio_bs=tasa_cambio,
                tipo=tipo,
                referencia=ref,
                estado="ACTIVO",
                tenant_id=current_user.tenant_id
            )
            db.add(nuevo_mov)
            total_monto_usd += monto_usd
            nuevos_count += 1

    cuenta.saldo_actual_usd = to_float(cuenta.saldo_actual_usd) + total_monto_usd
    db.commit()
    return {
        "ok": True,
        "count": len(movs),
        "conciliados": conciliados_count,
        "nuevos": nuevos_count,
        "message": f"Se procesaron {len(movs)} movimientos del extracto: {conciliados_count} conciliados contra registros existentes, {nuevos_count} nuevos sin coincidencia."
    }


@tesoreria_router.get("/movimientos-caja")
def movimientos_caja(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cuentas_caja = db.query(CuentaBancaria).filter(
        CuentaBancaria.activa == True,
        CuentaBancaria.banco.like("%Caja%"),
        CuentaBancaria.tenant_id == current_user.tenant_id
    ).all()
    cuenta_ids = [c.id for c in cuentas_caja]
    
    saldo_caja = sum(to_float(c.saldo_actual_usd) for c in cuentas_caja)
    
    movs = db.query(MovimientoBancario).filter(
        MovimientoBancario.cuenta_id.in_(cuenta_ids),
        MovimientoBancario.tenant_id == current_user.tenant_id
    ).order_by(MovimientoBancario.fecha.desc()).all()
    
    no_deducibles = 0.0
    soportes_count = 0
    now = datetime.now(timezone.utc)
    
    for m in movs:
        is_current_month = m.fecha.year == now.year and m.fecha.month == now.month
        if m.tipo == "EGRESO":
            if not m.referencia:
                if is_current_month:
                    no_deducibles += to_float(m.monto_usd)
        if m.referencia:
            soportes_count += 1
            
    soportes_pct = (soportes_count / len(movs) * 100.0) if len(movs) > 0 else 100.0
    
    return {
        "metricas": {
            "saldo_caja": f"${saldo_caja:,.2f}",
            "no_deducibles": f"${no_deducibles:,.2f}",
            "soportes_pct": f"{soportes_pct:.1f}%"
        },
        "movimientos": [
            {
                "id": m.id,
                "date": m.fecha.strftime("%d/%m/%Y") if m.fecha else "",
                "desc": m.concepto,
                "amount": f"{'+' if m.tipo == 'INGRESO' else '-'}${to_float(m.monto_usd):,.2f}",
                "support": "Factura" if m.referencia else "Sin Soporte",
                "fiscal": "Deducible" if m.referencia else "No Deducible",
                "fColor": "bg-green-100 text-green-700" if m.referencia else "bg-amber-100 text-amber-700",
                "hasImage": bool(m.referencia),
                "imageType": "file" if m.referencia else "none",
                "cuenta_nombre": m.cuenta.banco if m.cuenta else "N/A"
            }
            for m in movs
        ]
    }


@tesoreria_router.post("/movimientos-caja")
def registrar_movimiento_caja(body: dict, db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    cuenta_id = body.get("cuenta_id")
    concepto = body.get("concepto", "")
    monto_usd = float(body.get("monto_usd", 0.0))
    tipo = body.get("tipo", "INGRESO")
    referencia = body.get("referencia", "")
    tasa_cambio_bs = float(body.get("tasa_cambio_bs", 1.0))
    
    cuenta = db.query(CuentaBancaria).filter(CuentaBancaria.id == cuenta_id, CuentaBancaria.tenant_id == current_user.tenant_id).with_for_update().first()
    if not cuenta:
        raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        
    factor = 1.0 if tipo == "INGRESO" else -1.0
    cuenta.saldo_actual_usd = to_float(cuenta.saldo_actual_usd) + (factor * monto_usd)
    
    mov = MovimientoBancario(
        cuenta_id=cuenta_id,
        concepto=concepto,
        monto_usd=monto_usd,
        tasa_cambio_bs=tasa_cambio_bs,
        tipo=tipo,
        referencia=referencia,
        estado="ACTIVO",
        tenant_id=current_user.tenant_id
    )
    db.add(mov)
    db.commit()
    return {"ok": True, "id": mov.id}


@tesoreria_router.get("/inversiones/exportar")
def exportar_inversiones_excel(db: Session = Depends(get_db), current_user = Depends(get_current_user)):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    
    colocaciones = db.query(ColocacionInversion).filter(ColocacionInversion.tenant_id == current_user.tenant_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rendimiento de Inversiones"
    
    header_fill = PatternFill(start_color="0B5156", end_color="0B5156", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ["Nombre", "Plazo (días)", "Capital (Bs)", "Tasa Anual (%)", "Interés Ganado (Bs)", "Tasa Inicial (Bs/USD)", "Resultado Real (USD)", "Fecha Registro"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, c in enumerate(colocaciones, 2):
        capital = to_float(c.capital_bs)
        tasa = to_float(c.tasa_interes_anual)
        plazo = int(c.plazo_dias or 30)
        interes = capital * (tasa / 100) * (plazo / 365)
        tasa_inicial = to_float(c.tasa_cambio_inicial)
        resultado_usd = interes / tasa_inicial if tasa_inicial > 0 else 0
        
        ws.cell(row=row_idx, column=1, value=c.nombre)
        ws.cell(row=row_idx, column=2, value=plazo)
        ws.cell(row=row_idx, column=3, value=capital)
        ws.cell(row=row_idx, column=4, value=tasa)
        ws.cell(row=row_idx, column=5, value=round(interes, 2))
        ws.cell(row=row_idx, column=6, value=tasa_inicial)
        ws.cell(row=row_idx, column=7, value=round(resultado_usd, 2))
        ws.cell(row=row_idx, column=8, value=c.fecha_inicio.strftime("%Y-%m-%d") if c.fecha_inicio else "")
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rendimiento_inversiones.xlsx"}
    )
